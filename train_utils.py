from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
import requests
from requests import ConnectionError
import time
import threading
from threading import Thread
import re

import openpyxl

def write_to_excel(file_path, sheet_name, data):
    try:
        # Load the existing workbook or create a new one
        workbook = openpyxl.load_workbook(file_path)

        # Select the sheet by name or create a new one if it doesn't exist
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)

        # Append data to the sheet
        sheet.append(data)

        # Save the workbook
        workbook.save(file_path)
        print(f"Data written to {file_path} successfully.")

    except Exception as e:
        print(f"Error: {e}")

# Example usage


class SliderExample(QWidget):
    def __init__(self):
        super().__init__()

        self.init_ui()

    def init_ui(self):
        self.motor_current_state = 0
        self.motor_curr = 0
        self.total_curr = 0
        self.ir_state = 0
        self.total_current_state = 0
        self.pwm = 0
        self.state = 0
        self.frequency = 0.0001
        self.period = 1000
        self.initial_time = time.time()
        self.url = "http://192.168.1.7/"
        self.url2 = "http://192.168.1.7/pwm"
        self.url3 = "http://192.168.1.7/readings"


        self.setFont(QFont('Arial', 16))
        self.label1 = QLabel('PWM: 0', self)
        self.label1.setGeometry(000, 200, 100, 100)

        self.slider = QSlider(Qt.Horizontal, self)

        self.slider.setMinimum(0)
        self.slider.setMaximum(255)
        self.slider.valueChanged.connect(self.on_slider_change)
        self.slider.setGeometry(100, 200, 200, 100)

        layout = QVBoxLayout(self)
        self.setWindowTitle('Slider Example')
        self.setGeometry(000, 000, 1000, 400)

        self.button = QPushButton("Motor_Current", self)
        self.button.setGeometry(100, 000, 200, 200)
        self.button.setCheckable(True)
        self.button.clicked.connect(self.motor_current)
        self.button.setStyleSheet("background-color : red")
 
        self.button2 = QPushButton("Total_Current", self)
        self.button2.setGeometry(300, 000, 200, 200)
        self.button2.setCheckable(True)
        self.button2.clicked.connect(self.total_current)
        self.button2.setStyleSheet("background-color : red")

        self.button3 = QPushButton("IR", self)
        self.button3.setGeometry(500, 000, 200, 200)
        self.button3.setCheckable(True)
        self.button3.clicked.connect(self.ir_handle)
        self.button3.setStyleSheet("background-color : red")


        self.motor_data = QLabel("Hello, PyQt!", self)
        self.motor_data.setGeometry(700, 000, 200, 100)

        self.total_data = QLabel("Hello, PyQt!", self)
        self.total_data.setGeometry(700, 100, 200, 100)

        self.freq = QLabel(self)
        self.freq.setText('Frequency:')
        self.freq.setGeometry(900, 000, 200, 100)
        self.line = QLineEdit(self)
        self.line.setGeometry(900, 100, 200, 100)

        self.button3 = QPushButton("Run", self)
        self.button3.setGeometry(800, 200, 200, 100)
        self.button3.setCheckable(True)
        self.button3.clicked.connect(self.run_or_stop)
        self.button3.setStyleSheet("background-color : red")

        self.update()
        self.show()


    def motor_current(self):        
        # if button is checked
        if self.button.isChecked():
            self.button.setStyleSheet("background-color : green")
            self.motor_current_state=1
            print("motor current now 1")
            print(self.motor_current_state)
        # if it is unchecked
        else: 
            self.button.setStyleSheet("background-color : red")
            self.motor_data.setText("Motor button unpressed")
            self.motor_current_state=0
            print("motor current now 0")
            print(self.motor_current_state)

    def ir_handle(self):        
        # if button is checked
        if self.button.isChecked():
            self.button.setStyleSheet("background-color : green")
            self.ir_read=1
            print("Ir Reading now on")
        # if it is unchecked
        else: 
            self.button.setStyleSheet("background-color : red")
            self.ir_read=0
            print("IR reading now off")


    def run_or_stop(self):
        
        # if button is checked
        if self.button3.isChecked():
            self.button3.setStyleSheet("background-color : green")
            self.frequency = int(self.line.text())
            self.period = 1/self.frequency
            print(self.frequency)
            self.state=1
            
            reque()

        # if it is unchecked
        else:
            self.button3.setStyleSheet("background-color : red")
            self.state=0


    def total_current(self): 
        # if button is checked
        if self.button2.isChecked():
            self.button2.setStyleSheet("background-color : green")
            self.total_data.setText("Total Button pressed")
            self.total_current_state=1
            print("total current now 1")
            print(self.total_current_state)

        # if it is unchecked
        else:
            self.button2.setStyleSheet("background-color : red")
            self.total_data.setText("Total Button unpressed")
            self.total_current_state=0

    def on_slider_change(self):
        # write what happens on sliding the slider
        slider_value = self.slider.value()
        self.label1.setText(f'Slider Value: {slider_value}')
        self.pwm = slider_value
        send_request(self)


def send_request(qt_gui):
    try:
        data = {'pwm': str(qt_gui.pwm)}
        print(data)
        response = requests.post(qt_gui.url2, json=data)
        if response.status_code == 200:
            print("PWM = " + str(qt_gui.pwm))

        else:
            print(f"Request failed with status code: {response.status_code}")

    except requests.RequestException as e:
        print(f"Request failed: {e}")

def reque():
    global qtgui_obj
    tt = threading.Timer(qtgui_obj.period, reque)
    tt.start()   # run code every 5 seconds
    try:
        # Make a POST request with JSON data
        jso = {}
        jso['motor'] =  str(qtgui_obj.motor_current_state)
        jso['total'] =  str(qtgui_obj.total_current_state)
        jso['ir_state'] =  str(qtgui_obj.ir_state)

        response = requests.post(qtgui_obj.url3, json=jso)
        content = response.text
        #response = requests.get(url2)

        # Check if the request was successful (status code 200)
        if response.status_code == 200:
            aa = re.findall(r'\d+', response.text)
            qtgui_obj.motor_curr = aa[0]     
            qtgui_obj.total_curr = aa[1] 
        
            qtgui_obj.motor_data.setText(str(qtgui_obj.motor_curr))
            qtgui_obj.total_data.setText(str(qtgui_obj.total_curr))
            file_path = "C:/Users/Suhrida/Desktop/data.xlsx"
            sheet_name = "Sheet1"
            data_to_write = [time.time(), qtgui_obj.motor_curr, qtgui_obj.motor_curr, aa[2]]

            write_to_excel(file_path, sheet_name, data_to_write)
    
            if(qtgui_obj.state==0):
                tt.cancel()
        else:
            print(f"Request failed with status code: {response.status_code}")

    except requests.RequestException as e:
        print(f"Request failed: {e}")

def est_connection(qtgui):
    try:
        response = requests.get(qtgui.url)

        if response.status_code == 200:
            print("Connection established")
            print(response.text)
        else:
            print(f"Request failed with status code: {response.status_code}")

    except requests.RequestException as e:
        print(f"Request failed: {e}")



if __name__ == '__main__':
    
    app = QApplication([])
    qtgui_obj = SliderExample()
    est_connection(qtgui_obj)
    print(qtgui_obj.pwm)

    qtgui_obj.show()
    app.exec_()


